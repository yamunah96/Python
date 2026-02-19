# rows and columns
# module openpyxl

# pip (python package manger)- nstall other modules 
from openpyxl import Workbook
'''
# write data in excel
wb= Workbook()
ws= wb.active
ws.title= "Student Data"

ws['A1']="Name"
ws["B1"]="Score"
ws.append(["Anish",90])
ws.append(['Mira',85])

wb.save("students.xlsx")
'''



# Read data from existing excel file

from openpyxl import load_workbook

wb=load_workbook(r'C:\Users\Mamatha-Win10\Downloads\sales_data.xlsx')
sheet_name= wb['Sheet1']

for row in sheet_name.iter_rows(values_only=True):
    print(row)

