from openpyxl import Workbook, load_workbook
import os



class ExcelSheet:
    def __init__(self,excelsheetpath="School.xlsx"):
        self.filepath=excelsheetpath         # basic requirement
        self.sheetname="Students"

        self.create_excelsheet()

    def create_excelsheet(self):
        if not os.path.exists(self.filepath):
            new_workbook= Workbook()
            new_worksheet= new_workbook.active
            new_worksheet.title=self.sheetname
            new_worksheet.append(["Name","Grade","Roll_No"])
            new_workbook.save(self.filepath)
        else:
            print("excel sheet already exist")


    def loadExcelSheet(self):
        return load_workbook(self.filepath)
    

    def add_user(self,name,grade,roll_no):
        wb= self.loadExcelSheet()
        ws= wb[self.sheetname]
        ws.append([name,grade,roll_no])
        wb.save(self.filepath)
        return True



e1= ExcelSheet()
e1.add_user("Riya",10,"r1_123")